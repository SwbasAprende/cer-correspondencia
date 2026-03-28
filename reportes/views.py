from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.utils import timezone
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from correspondencia.models import Documento, Trazabilidad
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


@login_required
def reporte_lista(request):
    hoy   = timezone.now().date()
    anio  = int(request.GET.get('anio', hoy.year))
    mes   = request.GET.get('mes', '')

    docs = Documento.objects.filter(fecha_radicacion__year=anio)
    if mes:
        docs = docs.filter(fecha_radicacion__month=mes)

    # Estadísticas
    total        = docs.count()
    por_estado   = {e: docs.filter(estado=e).count()   for e, _ in Documento.Estado.choices}
    por_tipo     = {t: docs.filter(tipo=t).count()     for t, _ in Documento.Tipo.choices}
    por_prioridad= {p: docs.filter(prioridad=p).count() for p, _ in Documento.Prioridad.choices}
    vencidos     = docs.filter(fecha_limite__lt=hoy).exclude(estado__in=['respondido','archivado']).count()
    entrada      = docs.filter(flujo='entrada').count()
    salida       = docs.filter(flujo='salida').count()

    # Paginación
    docs_ordenados = docs.order_by('-fecha_radicacion').select_related('responsable', 'radicado_por')
    paginator = Paginator(docs_ordenados, 20)
    page = request.GET.get('page', 1)
    try:
        docs_page = paginator.page(page)
    except PageNotAnInteger:
        docs_page = paginator.page(1)
    except EmptyPage:
        docs_page = paginator.page(paginator.num_pages)

    # Años disponibles
    anios = list(range(2024, hoy.year + 2))
    meses = [
        ('1','Enero'),('2','Febrero'),('3','Marzo'),('4','Abril'),
        ('5','Mayo'),('6','Junio'),('7','Julio'),('8','Agosto'),
        ('9','Septiembre'),('10','Octubre'),('11','Noviembre'),('12','Diciembre'),
    ]

    return render(request, 'reportes/lista.html', {
        'docs':          docs_page,
        'page_obj':      docs_page,
        'paginator':     paginator,
        'pagina_actual': docs_page.number,
        'num_paginas':   paginator.num_pages,
        'total':         total,
        'por_estado':    por_estado,
        'por_tipo':      por_tipo,
        'por_prioridad': por_prioridad,
        'vencidos':      vencidos,
        'entrada':       entrada,
        'salida':        salida,
        'anio':          anio,
        'mes':           mes,
        'anios':         anios,
        'meses':         meses,
        'hoy':           hoy,
        'estados': Documento.Estado.choices,
    })


@login_required
def exportar_excel(request):
    hoy  = timezone.now().date()
    anio = int(request.GET.get('anio', hoy.year))
    mes  = request.GET.get('mes', '')

    docs = Documento.objects.filter(fecha_radicacion__year=anio)
    if mes:
        docs = docs.filter(fecha_radicacion__month=mes)
    docs = docs.order_by('-fecha_radicacion')

    # ── Crear libro Excel ──────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Correspondencia {anio}'

    # Colores institucionales
    VERDE  = 'FF1B4332'
    DORADO = 'FFD4A017'
    VERDE2 = 'FF2D6A4F'
    GRIS   = 'FFF8F9FA'
    BLANCO = 'FFFFFFFF'

    # ── Fila 1: Título institucional ───────────────────────────────────────
    ws.merge_cells('A1:L1')
    ws['A1'] = 'CENTRO DE ESTUDIOS REGIONALES — CER'
    ws['A1'].font      = Font(name='Calibri', bold=True, size=14, color=BLANCO)
    ws['A1'].fill      = PatternFill('solid', fgColor=VERDE)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # ── Fila 2: Subtítulo ──────────────────────────────────────────────────
    periodo = f'Año {anio}' + (f' — Mes {mes}' if mes else ' — Todos los meses')
    ws.merge_cells('A2:L2')
    ws['A2'] = f'Reporte de Correspondencia Institucional — {periodo}'
    ws['A2'].font      = Font(name='Calibri', size=11, color=BLANCO)
    ws['A2'].fill      = PatternFill('solid', fgColor=VERDE2)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 22

    # ── Fila 3: Generado ───────────────────────────────────────────────────
    ws.merge_cells('A3:L3')
    ws['A3'] = f'Generado el {hoy.strftime("%d/%m/%Y")} — Ley 594 de 2000 — Archivo General de la Nación'
    ws['A3'].font      = Font(name='Calibri', size=9, italic=True, color='FF6C757D')
    ws['A3'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[3].height = 16

    ws.append([])  # Fila 4 vacía

    # ── Fila 5: Encabezados de columna ─────────────────────────────────────
    encabezados = [
        'N°', 'Radicado', 'Tipo', 'Flujo', 'Asunto',
        'Remitente', 'Destinatario', 'Entidad',
        'Estado', 'Prioridad', 'Fecha Radicación', 'Fecha Límite'
    ]
    ws.append(encabezados)
    fila_enc = ws.max_row
    for col, _ in enumerate(encabezados, 1):
        celda = ws.cell(row=fila_enc, column=col)
        celda.font      = Font(name='Calibri', bold=True, color=BLANCO, size=10)
        celda.fill      = PatternFill('solid', fgColor=VERDE)
        celda.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        celda.border    = Border(
            bottom=Side(style='thin', color=DORADO),
            right=Side(style='thin', color='FFDDDDDD'),
        )
    ws.row_dimensions[fila_enc].height = 30

    # ── Filas de datos ─────────────────────────────────────────────────────
    colores_estado = {
        'recibido':   'FFDBEAFE',
        'en_tramite': 'FFFEF3C7',
        'respondido': 'FFD1FAE5',
        'archivado':  'FFF3F4F6',
    }
    colores_prioridad = {
        'normal':       'FFF3F4F6',
        'urgente':      'FFFEE2E2',
        'confidencial': 'FFEDE9FE',
    }

    for i, doc in enumerate(docs, 1):
        fila = [
            i,
            doc.radicado,
            doc.get_tipo_display(),
            doc.get_flujo_display(),
            doc.asunto,
            doc.remitente,
            doc.destinatario,
            doc.entidad or '—',
            doc.get_estado_display(),
            doc.get_prioridad_display(),
            doc.fecha_radicacion.strftime('%d/%m/%Y'),
            doc.fecha_limite.strftime('%d/%m/%Y') if doc.fecha_limite else '—',
        ]
        ws.append(fila)
        fila_actual = ws.max_row
        color_fondo = GRIS if i % 2 == 0 else BLANCO

        for col in range(1, 13):
            celda = ws.cell(row=fila_actual, column=col)
            celda.font      = Font(name='Calibri', size=9)
            celda.alignment = Alignment(vertical='center', wrap_text=True)
            celda.border    = Border(
                bottom=Side(style='hair', color='FFDDDDDD'),
                right=Side(style='hair',  color='FFDDDDDD'),
            )
            # Color especial para estado y prioridad
            if col == 9:
                celda.fill = PatternFill('solid', fgColor=colores_estado.get(doc.estado, BLANCO))
            elif col == 10:
                celda.fill = PatternFill('solid', fgColor=colores_prioridad.get(doc.prioridad, BLANCO))
            else:
                celda.fill = PatternFill('solid', fgColor=color_fondo)

        ws.row_dimensions[fila_actual].height = 20

    # ── Anchos de columna ──────────────────────────────────────────────────
    anchos = [5, 20, 12, 10, 40, 25, 25, 25, 14, 14, 16, 14]
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    # ── Hoja 2: Resumen estadístico ───────────────────────────────────────
    ws2 = wb.create_sheet('Resumen')
    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 15

    ws2.merge_cells('A1:B1')
    ws2['A1'] = 'RESUMEN ESTADÍSTICO'
    ws2['A1'].font      = Font(bold=True, size=12, color=BLANCO)
    ws2['A1'].fill      = PatternFill('solid', fgColor=VERDE)
    ws2['A1'].alignment = Alignment(horizontal='center')

    secciones = [
        ('POR ESTADO', [(l, docs.filter(estado=v).count()) for v, l in Documento.Estado.choices]),
        ('POR TIPO',   [(l, docs.filter(tipo=v).count())   for v, l in Documento.Tipo.choices]),
        ('POR FLUJO',  [('Entrada', docs.filter(flujo='entrada').count()),
                        ('Salida',  docs.filter(flujo='salida').count())]),
        ('POR PRIORIDAD', [(l, docs.filter(prioridad=v).count()) for v, l in Documento.Prioridad.choices]),
    ]

    fila_r = 2
    for titulo, datos in secciones:
        ws2.cell(row=fila_r, column=1, value=titulo).font = Font(bold=True, color=BLANCO)
        ws2.cell(row=fila_r, column=1).fill = PatternFill('solid', fgColor=VERDE2)
        ws2.cell(row=fila_r, column=2).fill = PatternFill('solid', fgColor=VERDE2)
        fila_r += 1
        for etiqueta, cantidad in datos:
            ws2.cell(row=fila_r, column=1, value=etiqueta).font = Font(size=10)
            ws2.cell(row=fila_r, column=2, value=cantidad).font = Font(size=10, bold=True)
            ws2.cell(row=fila_r, column=2).alignment = Alignment(horizontal='center')
            fila_r += 1
        ws2.cell(row=fila_r, column=1, value='TOTAL').font = Font(bold=True)
        ws2.cell(row=fila_r, column=2, value=docs.count()).font = Font(bold=True)
        fila_r += 2

    # ── Respuesta HTTP ─────────────────────────────────────────────────────
    nombre = f'reporte_CER_{anio}{"_mes"+mes if mes else ""}.xlsx'
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{nombre}"'
    wb.save(response)
    return response